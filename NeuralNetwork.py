from openpyxl import load_workbook
from openpyxl import Workbook

import tensorflow as tf
import numpy as np
import matplotlib.pyplot as plt
import math
import random
import time
from statistics import mean

_K=100
_K2=10

wb = load_workbook('C&H.xlsx')
sheet = wb.active

i=0;

yes_sample=[]
no_sample=[]

for row in sheet.iter_rows():
	if i>=1:
		size=len(row)

		state=0
		col=12

		info={}
		p=[]

		info['name']=row[0].value
		info['label']=row[1].value
		info['start']=row[2].value
		info['lip']=row[9].value

		while col<size and row[col].value is not None:
			if col>=12:
				price=row[col].value
				#print(price)
				p.append(price)
			col=col+1;

		if row[1].value==1:
			state=1
			yes_sample.append([info,p])
		else:
			no_sample.append([info,p])
		#print(state)

	i=i+1

def initSample( yes,no ):
	in_group=[]
	out_group=[]
	info_group=[]
	for sample in yes:
		total=len(sample[1])
		p=[]
		lip=sample[0]['lip']
		min2=min(sample[1])
		max2=max(sample[1])
		for i in range(0,_K+1):
			index=lip*(i)/(_K)
			index=math.floor(index)
			if (i/_K)==1:
				index=index-1
			price=sample[1][index]
			#if len(sample[1][index:math.floor(index+(i/_K))])>0:
				#price=sum(sample[1][index:math.floor(index+(i/_K))])/len(sample[1][index:math.floor(index+(i/_K))])
			price=(price-min2)/(max2-min2)
			p.append(price)
		for i in range(1,_K2+1):
			index=lip+(total-lip)*(i)/(_K2)
			index=math.floor(index)
			if (i/_K2)==1:
				index=index-1
			price=sample[1][index]
			price=(price-min2)/(max2-min2)
			p.append(price)
		p.append(lip/total)
		#p.append(total)
		in_group.append(p)
		out_group.append([1.0,0.0])
		info_group.append(sample[0])
	for sample in no:
		total=len(sample[1])
		p=[]
		lip=sample[0]['lip']
		min2=min(sample[1])
		max2=max(sample[1])
		for i in range(0,_K+1):
			index=lip*(i)/(_K)
			index=math.floor(index)
			if (i/_K)==1:
				index=index-1
			price=sample[1][index]
			#if len(sample[1][index:math.floor(index+(i/_K))])>0:
				#price=sum(sample[1][index:math.floor(index+(i/_K))])/len(sample[1][index:math.floor(index+(i/_K))])
			price=(price-min2)/(max2-min2)
			p.append(price)
		for i in range(1,_K2+1):
			index=lip+(total-lip)*(i)/(_K2)
			index=math.floor(index)
			if (i/_K2)==1:
				index=index-1
			price=sample[1][index]
			price=(price-min2)/(max2-min2)
			p.append(price)
		p.append(lip/total)
		#p.append(total)
		in_group.append(p)
		in_group.append(p)
		out_group.append([0.0,1.0])
		info_group.append(sample[0])
	
	combined = list(zip(in_group, out_group,info_group))
	random.shuffle(combined)

	in_group[:], out_group[:], info_group[:]= zip(*combined)

	return [in_group,out_group,info_group]

samples=initSample(yes_sample,no_sample)

#print(samples[0])

# Parameters
learning_rate = 0.1
num_steps = 2500
display_step = 100

# Network Parameters
n_hidden_1 = 16 # 1st layer number of neurons
n_hidden_2 = 16 # 2nd layer number of neurons

pivot=math.floor(4*len(samples[0])/5)

x = np.array(samples[0][0:pivot], dtype=np.float32)
y = np.array(samples[1][0:pivot], dtype=np.float32)

num_input = len(x[0]) 
num_classes = len(y[0])

# tf Graph input
X = tf.placeholder("float", [None, num_input])
Y = tf.placeholder("float", [None, num_classes])

# Store layers weight & bias
weights = {
    'h1': tf.Variable(tf.random_normal([num_input, n_hidden_1])),
    'h2': tf.Variable(tf.random_normal([n_hidden_1, n_hidden_2])),
    'out': tf.Variable(tf.random_normal([n_hidden_2, num_classes]))
}
biases = {
    'b1': tf.Variable(tf.random_normal([n_hidden_1])),
    'b2': tf.Variable(tf.random_normal([n_hidden_2])),
    'out': tf.Variable(tf.random_normal([num_classes]))
}

# Create model
def neural_net(x):
    layer_1 = tf.add(tf.matmul(x, weights['h1']), biases['b1'])
    layer_2 = tf.add(tf.matmul(layer_1, weights['h2']), biases['b2'])
    out_layer = tf.matmul(layer_2, weights['out']) + biases['out']
    return out_layer

# Construct model
logits = neural_net(X)
prediction = tf.nn.softmax(logits)

# Define loss and optimizer
loss_op = tf.reduce_mean(tf.nn.softmax_cross_entropy_with_logits(logits=logits, labels=Y))
optimizer = tf.train.AdamOptimizer(learning_rate=learning_rate)
train_op = optimizer.minimize(loss_op)

# Evaluate model
correct_pred = tf.equal(tf.argmax(prediction, 1), tf.argmax(Y, 1))
accuracy = tf.reduce_mean(tf.cast(correct_pred, tf.float32))

# Initialize the variables (i.e. assign their default value)
init = tf.global_variables_initializer()

#pivot=math.floor(len(samples)/2)

# Start training

def crossValidation(training,testing):

	for step in range(1, num_steps+1):

		output=sess.run(train_op, feed_dict={X: training[0], Y: training[1]})

		if step % display_step == 0 or step == 1:
			loss, acc = sess.run([loss_op, accuracy], feed_dict={X: x, Y: y})
			#print("Step " + str(step) + ", Minibatch Loss= {:.4f}".format(loss) + ", Training Accuracy= {:.3f}".format(acc))

	#print("Optimization Finished!")

	output=sess.run(accuracy, feed_dict={X: training[0], Y: training[1]})
	#print("Training Accuracy:", output)

	def CH_Test(x,y,info):

		output=sess.run(prediction, feed_dict={X: np.array(x, dtype=np.float32)})

		i=0
		i2=0
		acc=0

		nwb = Workbook()
		ws = nwb.active

		TP=0
		TN=0
		FP=0
		FN=0

		for y_p in output:
			state=y[i][0]==1
			state_p=(y_p[0]>=y_p[1])
			if state==state_p:
				acc=acc+1
			else:
				ws.cell(row=i2+1,column=1).value=str(state)
				ws.cell(row=i2+1,column=2).value=info[i]['name']
				ws.cell(row=i2+1,column=3).value=info[i]['label']
				ws.cell(row=i2+1,column=4).value=info[i]['start']
				for j in range(0,(_K+_K2+1)):
					ws.cell(row=i2+1,column=j+5).value=x[i][j]
				i2=i2+1
			
			if state_p==1 and state==1:
				TP=TP+1
			elif state_p==0 and state==0:
				TN=TN+1
			elif state_p==1 and state==0:
				FP=FP+1
			elif state_p==0 and state==1:
				FN=FN+1

			i=i+1

		nwb.save('log.xlsx')

		#return [acc/len(y),TP,TN,FP,FN]
		return [acc/len(y),(TP)/(TP+FN),(TN)/(TN+FP),(FP)/(TN+FP),(FN)/(TP+FN),(TP)/(TP+FP)]
	

	info=CH_Test(testing[0],testing[1],testing[2])
	#print("Testing Accuracy:",acc)

	return [output,info[0],info[1],info[2],info[3],info[4],info[5]]

with tf.Session() as sess:

	sess.run(init)

	avg_training_error=0
	avg_testing_error=0

	avg_RC=0
	avg_TNR=0
	avg_FPR=0
	avg_FNR=0
	avg_PS=0

	for k in range(1,5):

		sp = samples[:]

		pivot=math.floor(k*len(sp[0])/5)
		pivot2=math.floor((k+1)*len(sp[0])/5)

		spx_train=sp[0][0:pivot]+sp[0][pivot2+1:len(sp[0])-1]
		spy_train=sp[1][0:pivot]+sp[1][pivot2+1:len(sp[1])-1]

		x = np.array(spx_train, dtype=np.float32)
		y = np.array(spy_train, dtype=np.float32)

		x2 = np.array(sp[0][pivot+1:pivot2], dtype=np.float32)
		y2 = np.array(sp[1][pivot+1:pivot2], dtype=np.float32)
		z2 = sp[2][pivot+1:pivot2]

		training=[x,y]
		testing=[x2,y2,z2]
		result=crossValidation(training,testing)
		#print("Training Accuracy:",result[0])
		#print("Testing Accuracy:",result[1])
		avg_training_error=avg_training_error+result[0]
		avg_testing_error=avg_testing_error+result[1]

		avg_RC=avg_RC+result[2]
		avg_TNR=avg_TNR+result[3]
		avg_FPR=avg_FPR+result[4]
		avg_FNR=avg_FNR+result[5]
		avg_PS=avg_PS+result[6]

	print("*** Results ***")
	print("Average Training Accuracy:",avg_training_error/4)
	print("Average Testing Accuracy:",avg_testing_error/4)
	print("Average Testing Recall:",avg_RC/4)
	print("Average Testing TNR:",avg_TNR/4)
	#print("Average Testing FPR:",avg_FPR/4)
	#print("Average Testing FNR:",avg_FNR/4)
	#print("Average Testing Precision:",avg_PS/4)


#https://medium.com/@nonthakon/%E0%B8%A5%E0%B8%AD%E0%B8%87%E0%B9%83%E0%B8%8A%E0%B9%89-tensorflow-%E0%B8%97%E0%B8%B3-linear-regression-f9e05d734441
