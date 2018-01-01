from openpyxl import load_workbook
from openpyxl import Workbook

import numpy as np
import matplotlib.pyplot as plt

wb = load_workbook('C&H.xlsx')
sheet = wb.active

data={};

i=0;



for row in sheet.iter_rows():
	if i>=1:
		size=len(row)
		name=row[0].value
		col=12
		print(name)
		p=[]
		while col<size and row[col].value is not None:
			if col>=12:
				price=row[col].value
				#print(price)
				p.append(price)
			col=col+1;
		data[name]=p
		#print(p)
	if i>=1:
		pivot=[row[7].value,row[8].value,row[9].value,row[10].value,row[11].value];
		ans='NO'
		if row[1].value==1:
			ans='YES'
		plt.title('('+ans+') '+name+' from '+row[2].value.strftime("%B %d, %Y")+' to '+row[6].value.strftime("%B %d, %Y"))
		fig = plt.gcf()
		plt.plot(p)
		plt.plot(pivot,[row[13+row[7].value-1].value,row[13+row[8].value-1].value,row[13+row[9].value-1].value,row[13+row[10].value-1].value,row[13+row[11].value-1].value])
		plt.draw()
		plt.savefig('plot/sample'+str(i)+'.png')
		#plt.show()
		plt.close(fig) 
	i=i+1