from openpyxl import load_workbook
from openpyxl import Workbook

import math
import random

wb = load_workbook('Dataset.xlsx')
sheet = wb.active

stocklist={}
namelist=[]
datelist=[]
data = sheet['C1:YN2448']

for row in sheet.iter_rows(min_col=2,min_row=1, max_col=2, max_row=2448):
	d=row[0].value
	datelist.append(d)

datelist=datelist[::-1]

for col in sheet.iter_cols(min_col=3,min_row=1, max_col=664, max_row=2448):
	name=col[0].value
	#print(name)
	j=1
	namelist.append(name)
	stocklist[name]=[]
	temp=0
	for row in col:
		v=row.value
		stocklist[name].append(v)
		j=j+1
	stocklist[name].pop(0)
	stocklist[name]=stocklist[name][::-1]

		#if not isinstance(v, float):
		#	v=temp
		#temp=v
	
	t=len(stocklist[name])
	for k in range(1,t):
		v=stocklist[name][k]
		if not isinstance(v, float):
			stocklist[name][k]=(stocklist[name][k-1])

	stocklist[name]=list(filter(None.__ne__, stocklist[name]))

def zigzag( close,p ):

	t=len(close)
	pivot=1
	isUp = 0
	wave = []
	if t==0:
		return wave

	if close[pivot+1]>close[pivot]:
		isUp=1

	for k in range(1, t-1):
		#print(t,k+1,pivot)
		if isUp == 1 and close[k+1] > close[pivot]:
			pivot=k+1
		elif isUp == 0 and close[k+1] < close[pivot]:
			pivot=k+1
		elif isUp == 1 and (close[k+1]-close[pivot])/close[pivot] < -p:
			wave.append(pivot)
			#wave.append(close[pivot])
			pivot=k+1
			isUp = 0
		elif isUp == 0 and (close[k+1]-close[pivot])/close[pivot] > p:
			wave.append(pivot)
			#wave.append(close[pivot])
			pivot=k+1
			isUp = 1

	return wave

def findCH(close,wave,datelist):

	
	pstart=[]
	pend=[]
	t=len(wave)

	if t==0:
		return []

	#date=close.reset_index()['Date']


	t2=len(close)
	datelist=datelist[len(datelist)-t2:len(datelist)];

	state=1
	k=0

	p1=close[wave[0]]
	
	tmp2=close[wave[0]]
	d=datelist[wave[0]]
	st=wave[0]
	st2=wave[0]
	st3=wave[0]
	st4=wave[0]
	st5=wave[0]
	w=0
	w2=0
	w3=0
	w4=0
	w5=0
	p=[]

	while k<t-1:
		current=close[wave[k]]
		next=close[wave[k+1]]
		if state==1:

			p2=p1
			p3=p1
			p4=p1
			p5=p1

			if next<current and next<p2:
				p2=next
				d2=datelist[wave[k+1]-1]
				st2=wave[k+1]
				w2=k+1
				state=2
			elif next>=current:
				p1=next
				d=datelist[wave[k+1]-1]
				st=wave[k+1]
				w=k+1
		elif state==2:
			if next<current and next<p2:
				p2=next
				d2=datelist[wave[k+1]-1]
				st2=wave[k+1]
				w2=k+1
			elif next>=current:
				if next>1.1*p1:
					p1=next
					d=datelist[wave[k+1]-1]
					st=wave[k+1]
					w=k+1
					state=1
				elif next<=1.1*p1 and next>=0.9*p1:
					p3=next
					d3=datelist[wave[k+1]-1]
					st3=wave[k+1]
					w3=k+1
					state=3
		elif state==3:
			if next<=0.5*(p2+p3):
				p1=p3
				p2=p1
				p3=p1
				p4=p1
				p5=p1

				d=datelist[st3-1]
				st=st3
				w=k+1

				p2=next
				d2=datelist[wave[k+1]-1]
				st2=wave[k+1]
				w2=k+1
				state=2
			else:
				p4=next
				d4=datelist[wave[k+1]-1]
				st4=wave[k+1]
				w4=k+1
				state=4
		elif state==4:
			if next<current and next<p4:
				if next>p2 and next>=0.5*(p2+p3):
					p4=next
					d4=datelist[wave[k+1]-1]
					st4=wave[k+1]
					w4=k+1
				else:
					p1=p3
					p2=p1
					p3=p1
					p4=p1
					p5=p1
					d=datelist[st3-1]
					st=st3
					w=k+1

					p2=next
					d2=datelist[wave[k+1]-1]
					st2=wave[k+1]
					w2=k+1
					state=2
			elif next>=current and next>p3:
				p5=next
				d5=datelist[wave[k+1]-1]
				st5=wave[k+1]
				w5=k+1
				state=5
		elif state==5:
			
			#pstart.append(d)
			#pend.append(datelist[wave[k]-1])
			#pstart.append(st)
			#pend.append(wave[k]-1)

			k2=st4
		
			while k2<len(close) and close[st3]>=close[k2]:
				#print('ongoing')
				d5=datelist[st5]
				k2=k2+1
				st5=k2
				
			pivot=[]
			pivot.append(st)
			pivot.append(st2)
			pivot.append(st3)
			pivot.append(st4)
			pivot.append(st5)
			pivot.append(d)
			pivot.append(d2)
			pivot.append(d3)
			pivot.append(d4)
			pivot.append(d5)
			p.append(pivot)

			p1=p5
			p2=p1
			p3=p1
			p4=p1
			p5=p1
			d=datelist[st5-1]
			st=st5
			w=w5

			p2=next
			d2=datelist[wave[k+1]-1]
			st2=wave[k+1]
			w2=k+1
			state=2
		k=k+1		

	#p.append(pstart)
	#p.append(pend)

	return p

t=len(namelist)

total=0

patternlist={}

for k in range(1,t):
	name=namelist[k]
	print(name)
	w=zigzag(stocklist[name],0.05)
	print(w)
	p=findCH(stocklist[name],w,datelist)
	patternlist[name]=p
	print(patternlist[name])	
	total=total+len(p)

print(total)
print(len(namelist))

nwb = Workbook()
ws = nwb.active

k=1

#print(patternlist)
datalist=[]

l=1
ws.cell(row=l,column=1).value='name'
ws.cell(row=l,column=2).value='label'
ws.cell(row=l,column=3).value='d1'
ws.cell(row=l,column=4).value='d2'
ws.cell(row=l,column=5).value='d3'
ws.cell(row=l,column=6).value='d4'
ws.cell(row=l,column=7).value='d5'
ws.cell(row=l,column=8).value='p1'
ws.cell(row=l,column=9).value='p2'
ws.cell(row=l,column=10).value='p3'
ws.cell(row=l,column=11).value='p4'
ws.cell(row=l,column=12).value='p5'
ws.cell(row=l,column=13).value='data'
l=2

for p in patternlist:
	total=len(patternlist[p])
	#print(patternlist[p])
	
	for q in range(0,total):
		k=1
		start=patternlist[p][q][0]
		end=patternlist[p][q][4]

		ws.cell(row=l,column=1).value=p
		ws.cell(row=l,column=2).value='?'
		ws.cell(row=l,column=3).value=patternlist[p][q][5]
		ws.cell(row=l,column=4).value=patternlist[p][q][6]
		ws.cell(row=l,column=5).value=patternlist[p][q][7]
		ws.cell(row=l,column=6).value=patternlist[p][q][8]
		ws.cell(row=l,column=7).value=patternlist[p][q][9]
		ws.cell(row=l,column=8).value=patternlist[p][q][0]-patternlist[p][q][0]
		ws.cell(row=l,column=9).value=patternlist[p][q][1]-patternlist[p][q][0]
		ws.cell(row=l,column=10).value=patternlist[p][q][2]-patternlist[p][q][0]
		ws.cell(row=l,column=11).value=patternlist[p][q][3]-patternlist[p][q][0]
		ws.cell(row=l,column=12).value=patternlist[p][q][4]-patternlist[p][q][0]

		k=13
		if patternlist[p][q][4]-patternlist[p][q][0]<=1000:
		#if 1==1:
			for r in range(start,end+1):
			#datalist.append(stocklist[p][start:end+1])
				ws.cell(row=l,column=k).value=stocklist[p][r]
				#ws.append(p)
				#ws.append(stocklist[p][start:end+1])
				k=k+1
			l=l+1

nwb.save('C&H.xlsx')
