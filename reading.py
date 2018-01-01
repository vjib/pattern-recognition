from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook('stocks.xlsx')
sheet = wb.active

namelist=[]

for i in range(1,669):
	name=sheet['A'+ str(i)].value
	name=name+".BK"
	namelist.append(name)

print(namelist)


nwb = Workbook()
ws = nwb.create_sheet("Data")

#ws.cell(row=1,column=1).value='date'
k=2

for name in namelist:
	print(name)
	ws.cell(row=1,column=k).value=name
#	w=zigzag(close[name],0.1)
	#print(w)
#	p=findCH(close[name],w)
#	print(p)
	k=k+1
	
nwb.save('dataset_'+str(pstart)+'_'+str(pend)+'.xlsx')